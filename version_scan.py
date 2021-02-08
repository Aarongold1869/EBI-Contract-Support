import os
import glob
import openpyxl 
from datetime import datetime

def check_version():
    invoices = glob.glob('**/*.xlsm', recursive=True)
    version_table = {}
    for invoice in invoices: # iterate through files in list
        wb = openpyxl.load_workbook(invoice, keep_vba=True) # open workbook
        if 'Task Order' in wb.sheetnames: # check if wb contains Task Order Sheet
            ws = wb['Task Order'] # set Task Order to active sheet
            ws.protection.password = 'ftainv'
            ws.protection.disable() # unprotect sheet

            version = ws['S3'].value
            version_table[invoice] = version

            ws.protection.enable()
            wb.save(invoice) # save workbook
            wb.close()

    print(version_table)


check_version()