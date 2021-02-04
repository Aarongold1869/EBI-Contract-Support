##############################
# This program programatically updates all invoice package workbooks 
# within a directory to the most recent VBA project version. 
##############################

import os
import openpyxl 
import glob
from zipfile import ZipFile
from datetime import datetime

# input file to copy vba from (user input)
# input directory you want to update (user input)
# extract newest vba file from most recent workbook template
# get version number from most recent workbook template
# get list of all files to update using glob
# iterate through list
    # change file to zip
    # remove old vba file
    # paste new copied vba file
    # convert file back to xlsm
    # open workbook and update version (paste from template) using openpyxl
    # save and close workbook using openpyxl
# create csv detailing which workbooks were updated, date updated, and which version they are running

# get newest vbaProject binary file form workbook
# extract version number from workbook
def get_updated_vba_project(newest_version):
    wb = openpyxl.load_workbook(newest_version, keep_vba=True) # open workbook
    if 'Task Order' in wb.sheetnames: # check if wb contains Task Order Sheet
        ws = wb['Task Order'] # set Task Order to active sheet
        ws.protection.password = 'ftainv'
        ws.protection.disable() # unprotect sheet
        version = ws['S3'].value
        ws.protection.enable()
        wb.save(newest_version) # save workbook
        wb.close()

    zip_file_name = newest_version.split('.zip')[0] +'.zip'
    os.rename(newest_version, zip_file_name)
    vba_file = 'xl/vbaProject.bin'
    with ZipFile(zip_file_name, 'r') as zipObj:
        zipObj.extract(vba_file)
    os.rename(zip_file_name, newest_version)

    cwd = os.getcwd()
    vba_path = cwd + '\\' + vba_file

    return([version, vba_path])

# update each workbook in directory with new version number and new vba binary file 
def update_invoices(version, vba_path, directory_to_update):
    print(version,vba_path)
    invoices = glob.glob(directory_to_update + '/**/*.xlsm', recursive=True)
    for invoice in invoices:
        wb = openpyxl.load_workbook(invoice, keep_vba=True) # open workbook
        if 'Task Order' in wb.sheetnames: # check if wb contains Task Order Sheet
            ws = wb['Task Order'] # set Task Order to active sheet
            ws.protection.password = 'ftainv'
            ws.protection.disable() # unprotect sheet
            if ws['S3'].value == version:
                ws.protection.enable()
                wb.save(invoice) # save workbook
                wb.close()
                continue
            else:
                ws['S3'].value = version
                ws.protection.enable()
                wb.save(invoice) # save workbook
                wb.close()
        else:
            print('task order sheet not found in: ', invoice)
            continue

        # convert file to zip 
        zip_file_name = invoice.split('.')[0] +'.zip'
        os.rename(invoice, zip_file_name + ' ')

        # delete old vba binary file from zip 
        zin = ZipFile (zip_file_name, 'r')
        zout = ZipFile ('temp.zip', 'w')
        for item in zin.infolist():
            buffer = zin.read(item.filename)
            if (item.filename == 'xl/vbaProject.bin'):
                continue
            else:
                zout.writestr(item, buffer)
        zout.close()
        zin.close()

        # add new vba binary file to zip 
        with ZipFile('temp.zip', 'a') as zipf:
            zipf.write(vba_path, 'xl/vbaProject.bin')

        # convert file back to xlsm
        os.rename('temp.zip', invoice)


# get newest vba project and directory to update from user
# call update function
def main():
    run = True
    while run == True:
        newest_version = input("Most recent Workbook Version (paste complete file path here): ")
        isfile = os.path.isfile(newest_version)
        if isfile == False:
            print('enter a valid macro enabled workbook file path. press ctrl + c to exit.')
            continue
        while run == True:
            directory_to_update = input("Enter directory you wish to update (paste folder path here): ")
            isdir = os.path.isdir(directory_to_update)  
            if isdir == False:
                print('directory not found. press ctrl + c to exit.')
                continue
            else:
                run = False
    data = get_updated_vba_project(newest_version)
    version = data[0]
    vba_path = data[1]
    update_invoices(version, vba_path, directory_to_update)

if __name__ == "__main__":
    main()
